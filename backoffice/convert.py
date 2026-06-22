#!/usr/bin/env python3
"""
Excel → JSON Converter for Priao VIP Catalog
=============================================

แปลงไฟล์ Excel template "ไฟล์อัพโหลดสินค้าเข้า Catalog Wholesale"
เป็นไฟล์ JSON 9 ไฟล์ (C01.json - C09.json) ที่ใช้ใน data/

USAGE:
    python convert.py <excel_file> [--out <output_dir>]

EXAMPLE:
    python convert.py products.xlsx --out ../data/

REQUIREMENTS:
    pip install openpyxl
"""

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: ต้องติดตั้ง openpyxl ก่อน → pip install openpyxl")
    sys.exit(1)


# ----------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------

# Mapping: หมวดสินค้า (Excel) → CXX (JSON filename)
CATEGORY_MAP = {
    "เครื่องสำอาง": "C01",
    "ผลิตภัณฑ์ดูแลผิวหน้า": "C02",
    "ผลิตภัณฑ์ดูแลผิวกาย": "C03",
    "ผลิตภัณฑ์ดูแลเส้นผม": "C04",
    "น้ำหอม": "C05",
    "อุปกรณ์เพื่อความงาม": "C06",
    "อาหารเสริม": "C07",
    "คอนซูเมอร์": "C08",
    "แฟชั่น&ไลฟ์สไตล์": "C09",
}

# Mapping: PSaleStatus → flag (last element in JSON row)
STATUS_FLAG_MAP = {"A": 1, "New": 2, "B": 3}

# Excel column index (1-based, row 3 is header, row 4+ is data)
# Schema updated 2026-06-19: drop suggest_retail, change promo schema
COL = {
    "category":    1,   # หมวดสินค้า
    "subcategory": 2,   # หมวดย่อยสินค้า
    "barcode":     3,   # รหัสสินค้า
    "name":        4,   # ชื่อสินค้า
    "pack_qty":    5,   # จำนวน/หน่วย
    "unit":        6,   # หน่วย
    "brand":       7,   # แบรนด์
    "psale":       8,   # PSaleStatus
    "tag":         9,   # Tag
    "price":       10,  # price_wholesale
    "stock":       11,  # OnStockFG
    "image":       12,  # ลิงก์รูปภาพ (hyperlink)
    # promo columns (optional — empty if no promo)
    "promo_type":  13,  # "step price" / "flash" / (empty)
    "promo_label": 14,  # "ซื้อสินค้า 6 หน่วยขึ้นไป" / "⚡ FLASH SALE" / auto
    "promo_price": 15,  # ราคาพิเศษเมื่อเข้าเงื่อนไข (ตัวเลข)
    "condition":   16,  # เงื่อนไข (sentence — for validation only)
}

# Map promo_type → min_qty threshold (single source of truth)
PROMO_MIN_QTY = {
    "step_price": 6,   # ซื้อ 6 ชิ้นขึ้นไป → ใช้ promo_price ทั้งหมด
    "flash":      1,   # ซื้อ 1 ชิ้นขึ้นไป → ใช้ promo_price ทันที
}

# Default labels (เมื่อ Excel ปล่อย col 14 ว่าง — auto-generate)
PROMO_DEFAULT_LABELS = {
    "step_price": "ซื้อสินค้า 6 หน่วยขึ้นไป",
    "flash":      "⚡ FLASH SALE",
}

DATA_START_ROW = 4   # หัวตาราง row 3, data เริ่ม row 4


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def s(v):
    """Safe string conversion (None → empty)."""
    return "" if v is None else str(v).strip()


def to_int(v, default=0):
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def to_float(v, default=0):
    if v is None or v == "":
        return default
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return default


# ----------------------------------------------------------------------
# Main converter
# ----------------------------------------------------------------------
def convert(xlsx_path: Path, out_dir: Path, brands_dir: Path = None):
    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title} (rows: {ws.max_row}, cols: {ws.max_column})")

    # Validate headers (row 3)
    headers = [s(c.value) for c in ws[3]]
    expected = ["หมวดสินค้า", "หมวดย่อยสินค้า", "รหัสสินค้า", "ชื่อสินค้า",
                "จำนวน/หน่วย", "หน่วย", "แบรนด์", "PSaleStatus", "Tag",
                "price_wholesale", "OnStockFG", "ลิงก์รูปภาพ"]
    for i, exp in enumerate(expected):
        if i < len(headers) and headers[i] != exp:
            print(f"  WARN: column {i+1} header '{headers[i]}' ≠ expected '{exp}'")

    # Existing brand asset slugs (for warning if missing)
    existing_brand_slugs = set()
    if brands_dir and brands_dir.exists():
        for p in brands_dir.iterdir():
            if p.is_file():
                existing_brand_slugs.add(p.stem.upper())

    # Process rows
    by_category = defaultdict(list)
    warnings = []
    skipped = 0
    auto_fixed_stock = 0
    seen_barcodes = set()
    new_brands = set()

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        cat_name = s(ws.cell(row_idx, COL["category"]).value)
        if not cat_name or cat_name not in CATEGORY_MAP:
            # Skip subtotal rows / unknown categories
            if cat_name and "รวมทั้งหมด" not in cat_name:
                warnings.append(f"Row {row_idx}: ข้ามแถวที่หมวดสินค้า='{cat_name}' (ไม่อยู่ใน mapping)")
            skipped += 1
            continue

        cat_code = CATEGORY_MAP[cat_name]

        barcode    = s(ws.cell(row_idx, COL["barcode"]).value)
        name       = s(ws.cell(row_idx, COL["name"]).value)
        subcat     = s(ws.cell(row_idx, COL["subcategory"]).value)
        pack_qty   = to_int(ws.cell(row_idx, COL["pack_qty"]).value, 1)
        unit       = s(ws.cell(row_idx, COL["unit"]).value) or "ชิ้น"
        brand      = s(ws.cell(row_idx, COL["brand"]).value)
        psale      = s(ws.cell(row_idx, COL["psale"]).value)
        tag        = s(ws.cell(row_idx, COL["tag"]).value)
        price      = to_float(ws.cell(row_idx, COL["price"]).value, 0)
        stock      = to_int(ws.cell(row_idx, COL["stock"]).value, 0)

        # Image: ดึงจาก hyperlink ของ cell (ไม่ใช่ value ที่เป็น "📷 ดูรูป")
        img_cell = ws.cell(row_idx, COL["image"])
        image_url = ""
        if img_cell.hyperlink and img_cell.hyperlink.target:
            image_url = img_cell.hyperlink.target
        elif img_cell.value and s(img_cell.value).startswith("http"):
            # บางครั้งทีมอาจ paste URL ตรงๆ
            image_url = s(img_cell.value)

        # Promo fields (optional)
        # Normalize promo_type: "step price"/"flash" → "step_price"/"flash" (snake_case)
        promo_type_raw = s(ws.cell(row_idx, COL["promo_type"]).value).lower().strip()
        promo_type_norm = promo_type_raw.replace(" ", "_")
        promo_type = promo_type_norm if promo_type_norm in PROMO_MIN_QTY else ""

        # promo_label: trim + fallback to default if empty
        promo_label = s(ws.cell(row_idx, COL["promo_label"]).value).strip()
        if promo_type and not promo_label:
            promo_label = PROMO_DEFAULT_LABELS.get(promo_type, "")

        # promo_price: ราคาพิเศษ (ตัวเลข)
        promo_price = to_float(ws.cell(row_idx, COL["promo_price"]).value, 0)

        # promo_min_qty: derive from promo_type (single source of truth)
        promo_min_qty = PROMO_MIN_QTY.get(promo_type, 0)

        # Validation: เงื่อนไข (col P) ตรงกับ promo_type ที่ derive หรือไม่
        cond_text = s(ws.cell(row_idx, COL["condition"]).value).strip()
        if promo_type and cond_text:
            m = re.search(r"\d+", cond_text)
            cond_qty = int(m.group(0)) if m else 0
            if cond_qty and cond_qty != promo_min_qty:
                warnings.append(
                    f"Row {row_idx}: barcode={barcode} promo_type={promo_type} qty mismatch — "
                    f"derive={promo_min_qty} from promo_type but col P says qty={cond_qty}"
                )

        # Sanity check: ต้องมี promo_price ถ้ามี promo_type
        if promo_type and not promo_price:
            warnings.append(f"Row {row_idx}: barcode={barcode} promo_type='{promo_type}' but promo_price empty → skip promo")
            promo_type = ""
            promo_label = ""
            promo_min_qty = 0

        # Validation
        if not barcode:
            warnings.append(f"Row {row_idx}: ไม่มีรหัสสินค้า ข้าม")
            skipped += 1
            continue

        if barcode in seen_barcodes:
            warnings.append(f"Row {row_idx}: barcode '{barcode}' ซ้ำ (เก็บอันแรก)")
            continue
        seen_barcodes.add(barcode)

        # Auto-override Tag เมื่อ stock = 0
        if stock <= 0 and tag != "สินค้าหมดชั่วคราว":
            if tag:
                warnings.append(f"Row {row_idx}: barcode={barcode} stock=0 but Tag='{tag}' → auto-override")
            tag = "สินค้าหมดชั่วคราว"
            auto_fixed_stock += 1

        # Map PSaleStatus → flag
        flag = STATUS_FLAG_MAP.get(psale)
        if flag is None:
            warnings.append(f"Row {row_idx}: PSaleStatus='{psale}' ไม่รู้จัก (ใช้ flag=3 default)")
            flag = 3

        # Brand check
        if brand:
            brand_first = brand.split()[0] if " " in brand else brand
            slug = re.sub(r"[^A-Za-z0-9]", "", brand_first).upper()
            if existing_brand_slugs and slug not in existing_brand_slugs:
                new_brands.add((brand, slug))

        # Build pack_label "1 / แท่ง"
        pack_label = f"{pack_qty} / {unit}"

        # Build product row
        # Base: 11 elements (cols 0-10) — required for all products
        # Promo: 4 elements (cols 11-14) — only when promo_type is set
        # Total: 11 (no promo) or 15 (with promo)
        product = [
            barcode,       # 0
            name,          # 1
            subcat,        # 2
            tag,           # 3
            price,         # 4
            stock,         # 5
            image_url,     # 6
            brand,         # 7
            pack_qty,      # 8
            pack_label,    # 9
            flag,          # 10
        ]
        # Promo block (4 elements) — only if valid promo
        if promo_type:
            product += [
                promo_type,    # 11 — "step_price" | "flash"
                promo_label,   # 12 — ribbon text
                promo_price,   # 13 — discounted unit price
                promo_min_qty, # 14 — qty threshold (6 or 1)
            ]

        by_category[cat_code].append(product)

    # ------------------------------------------------------------------
    # Write JSON outputs
    # ------------------------------------------------------------------
    out_dir.mkdir(parents=True, exist_ok=True)
    print("\nWriting JSON files:")
    for cat_code in sorted(CATEGORY_MAP.values()):
        items = by_category.get(cat_code, [])
        out_path = out_dir / f"{cat_code}.json"
        with out_path.open("w", encoding="utf-8") as f:
            f.write("[\n")
            for i, item in enumerate(items):
                line = json.dumps(item, ensure_ascii=False)
                f.write(line + ("," if i < len(items) - 1 else "") + "\n")
            f.write("]\n")
        print(f"  {cat_code}.json: {len(items):,} รายการ")

    # Update data/index.json
    idx_path = out_dir / "index.json"
    idx_path.write_text(
        json.dumps({"categories": sorted(CATEGORY_MAP.values())}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  index.json: {len(CATEGORY_MAP)} categories")

    # ------------------------------------------------------------------
    # Report
    # ------------------------------------------------------------------
    report_lines = [
        f"# Conversion Report",
        f"",
        f"**Source:** `{xlsx_path.name}`",
        f"**Output:** `{out_dir}/`",
        f"",
        f"## Summary",
        f"- รายการทั้งหมด: {sum(len(v) for v in by_category.values()):,}",
        f"- ข้าม (subtotal/invalid): {skipped:,}",
        f"- Auto-fix stock=0 → Tag หมดชั่วคราว: {auto_fixed_stock:,}",
        f"- คำเตือน: {len(warnings):,}",
        f"",
        f"## By category",
    ]
    for cat_name, cat_code in sorted(CATEGORY_MAP.items(), key=lambda kv: kv[1]):
        cnt = len(by_category.get(cat_code, []))
        report_lines.append(f"- **{cat_name}** ({cat_code}): {cnt:,} รายการ")

    if new_brands:
        report_lines += ["", "## ⚠️ Brands ที่อาจยังไม่มีไฟล์รูปใน assets/brands/"]
        for brand, slug in sorted(new_brands):
            report_lines.append(f"- `{brand}` → คาดว่าจะหาไฟล์ชื่อ `{slug}.png` หรือ `{slug}.jpg`")

    if warnings:
        report_lines += ["", f"## คำเตือน ({len(warnings)} รายการ)"]
        for w in warnings[:50]:
            report_lines.append(f"- {w}")
        if len(warnings) > 50:
            report_lines.append(f"- ... และอีก {len(warnings) - 50} รายการ")

    report_path = out_dir.parent / "conversion_report.md"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")
    print(f"\nReport saved: {report_path}")

    print(f"\n✓ DONE — {sum(len(v) for v in by_category.values()):,} รายการแปลงเรียบร้อย")
    if new_brands:
        print(f"  ⚠ {len(new_brands)} brand ใหม่ที่ยังไม่มีไฟล์รูป (ดู report)")
    if warnings:
        print(f"  ⚠ {len(warnings)} คำเตือน (ดู report)")


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description="แปลง Excel template เป็น JSON สำหรับ data/ ของ Priao VIP Catalog",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("excel", type=Path, help="ไฟล์ Excel ต้นฉบับ (.xlsx)")
    ap.add_argument("--out", type=Path, default=Path("../data"),
                    help="โฟลเดอร์ output (default: ../data)")
    ap.add_argument("--brands-dir", type=Path, default=Path("../assets/brands"),
                    help="โฟลเดอร์ brand assets เพื่อเช็ค brand ใหม่ (default: ../assets/brands)")
    args = ap.parse_args()

    if not args.excel.exists():
        print(f"ERROR: ไม่พบไฟล์ {args.excel}")
        sys.exit(1)

    convert(args.excel, args.out, args.brands_dir)


if __name__ == "__main__":
    main()
